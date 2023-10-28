import openpyxl

book = openpyxl.load_workbook("../../data/unit 1/SOWC 2014 Stat Tables_Table 9.xlsx")
sheet = book["Table 9 "]

data = {}

# Assuming data starts from 15th row
for row in sheet.iter_rows(min_row=15, max_row=sheet.max_row, values_only=True):

    country = row[1]

    data[country] = {
        'child_labor': {
            'total': [row[4], row[5]],
            'male': [row[6], row[7]],
            'female': [row[8], row[9]],
        },
        'child_marriage': {
            'married_by_15': [row[10], row[11]],
            'married_by_18': [row[12], row[13]],
        }
    }

    if country == "Zimbabwe":
        break

import pprint
pprint.pprint(data)
